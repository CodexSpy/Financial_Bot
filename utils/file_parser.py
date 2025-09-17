from langchain.document_loaders import PyMuPDFLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter
import openpyxl
from langchain.schema import Document


# pdf file processor

def load_pdf(file_path):
    loader = PyMuPDFLoader(file_path)
    documents = loader.load()

    texts = [doc.page_content for doc in documents]
    text_splitter=RecursiveCharacterTextSplitter(chunk_size=700, chunk_overlap=200)
    split_texts = text_splitter.split_text('\n'.join(texts))

    final_doc = [
        Document(page_content=chunk, metadata={'source': file_path})
        for chunk in split_texts
    ]

    return final_doc

# excel files processor

def load_xlsx(file_path):
    workbook = openpyxl.load_workbook(file_path)
    documents = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        content = []

        for row in sheet.iter_rows(values_only=True):
            row_data = [str(cell) if cell is not None else "" for cell in row]
            if any(row_data):
                content.append(" | ".join(row_data))

        if content:
            document_text = "\n".join(content)
            doc = Document(
                page_content=document_text,
                metadata={"source": file_path, "sheet": sheet_name}
            )
            documents.append(doc)

    text_splitter = RecursiveCharacterTextSplitter(chunk_size=700, chunk_overlap=200)
    final_doc = text_splitter.split_documents(documents)

    return final_doc